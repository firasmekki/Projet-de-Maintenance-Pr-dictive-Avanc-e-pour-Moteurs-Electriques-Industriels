"""Generate downloadable export files (PDF, CSV, XLSX, JSON) — industrial quality."""
from __future__ import annotations

import csv
import io
import json
from datetime import datetime
from typing import Any


_FAULT_FR = {
    "No Fault":          "Aucun Défaut",
    "Early Degradation": "Dégradation Précoce",
    "Bearing Wear":      "Usure des Roulements",
    "Misalignment":      "Désalignement",
    "Unbalance":         "Déséquilibre Rotor",
    "Rotor Fault":       "Défaut Rotor",
    "Insulation Fault":  "Défaut d'Isolation",
    "Overload":          "Surcharge",
}

_SEV_FR = {"LOW": "FAIBLE", "MEDIUM": "MOYEN", "HIGH": "ÉLEVÉ", "CRITICAL": "CRITIQUE"}
_TREND_FR = {"RISING": "↑ HAUSSE", "FALLING": "↓ BAISSE", "STABLE": "→ STABLE"}
_STATUS_FR = {"Healthy": "Sain", "Warning": "Avertissement", "Critical": "Critique"}


class ExportService:

    # ── JSON ─────────────────────────────────────────────────────────────

    def to_json(self, report: Any) -> bytes:
        analysis = json.loads(report.analysis_json) if report.analysis_json else {}
        payload = {
            "report_id":      str(report.id),
            "filename":       report.filename,
            "created_at":     report.created_at.isoformat(),
            "analyzed_at":    report.analyzed_at.isoformat() if report.analyzed_at else None,
            "row_count":      report.row_count,
            "column_count":   report.column_count,
            "missing_values": report.missing_values,
            "quality_score":  report.quality_score,
            "columns":        json.loads(report.columns_json),
            "analysis":       analysis,
            "ai_narrative":   report.ai_narrative,
        }
        return json.dumps(payload, indent=2, ensure_ascii=False).encode("utf-8")

    # ── CSV ──────────────────────────────────────────────────────────────

    def to_csv(self, report: Any) -> bytes:
        if not report.dataset_json:
            return b"No dataset available"
        records = json.loads(report.dataset_json)
        if not records:
            return b"Empty dataset"
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=list(records[0].keys()))
        writer.writeheader()
        writer.writerows(records)
        return buf.getvalue().encode("utf-8")

    # ── Excel ─────────────────────────────────────────────────────────────

    def to_xlsx(self, report: Any) -> bytes:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb  = openpyxl.Workbook()
        analysis = json.loads(report.analysis_json) if report.analysis_json else {}

        # Sheet 1: Dataset
        ws_data = wb.active
        ws_data.title = "Dataset"
        if report.dataset_json:
            records = json.loads(report.dataset_json)
            if records:
                headers = list(records[0].keys())
                self._write_header_row(ws_data, headers)
                for row_data in records:
                    ws_data.append([row_data.get(h) for h in headers])
                for ci in range(1, len(headers) + 1):
                    ws_data.column_dimensions[get_column_letter(ci)].width = 16

        # Sheet 2: Analysis
        ws_res = wb.create_sheet("Résultats Analyse")
        risk = analysis.get("risk", {}); anomaly = analysis.get("anomaly", {})
        mp   = analysis.get("motor_profile") or {}
        rul  = analysis.get("rul", {})
        rows = [
            ["Indicateur",          "Valeur"],
            ["Fichier analysé",     report.filename],
            ["Date analyse",        report.analyzed_at.strftime("%Y-%m-%d %H:%M") if report.analyzed_at else "—"],
        ]
        if mp:
            rows += [
                ["Moteur",              mp.get("name", "—")],
                ["Fabricant",           mp.get("manufacturer", "—")],
                ["Puissance nominale",  f"{mp.get('nominal_power_kw', '—')} kW"],
                ["Courant nominal",     f"{mp.get('nominal_current_a', '—')} A"],
                ["Vitesse nominale",    f"{mp.get('nominal_speed_rpm', '—')} tr/min"],
                ["Classe isolation",    mp.get("insulation_class", "—")],
                ["Classe efficacité",   mp.get("efficiency_class", "—")],
            ]
        rows += [
            ["", ""],
            ["Score de Santé",      f"{analysis.get('health_score', '—')}/100"],
            ["Statut",              _STATUS_FR.get(analysis.get("health_status", ""), analysis.get("health_status", "—"))],
            ["Défaut Détecté",      _FAULT_FR.get(analysis.get("fault", ""), analysis.get("fault", "—"))],
            ["Sévérité",            _SEV_FR.get(analysis.get("severity", ""), "—")],
            ["Confiance",           f"{analysis.get('confidence', 0)}%"],
            ["Zone ISO 10816",      analysis.get("iso_zone", "—")],
            ["Niveau de Risque",    _SEV_FR.get(analysis.get("risk_level", ""), "—")],
            ["Risque 7 jours",      f"{risk.get('days_7', 0):.1f}%"],
            ["Risque 30 jours",     f"{risk.get('days_30', 0):.1f}%"],
            ["RUL estimé",          rul.get("value", "—")],
            ["Anomalies",           f"{anomaly.get('count', 0)} ({anomaly.get('percentage', 0):.1f}%)"],
        ]
        self._write_header_row(ws_res, ["Indicateur", "Valeur"])
        for r in rows[1:]:
            ws_res.append(r)
        ws_res.column_dimensions["A"].width = 26
        ws_res.column_dimensions["B"].width = 50

        # Sheet 3: Statistics
        ws_stat = wb.create_sheet("Statistiques")
        stats = analysis.get("statistics", {})
        self._write_header_row(ws_stat, ["Capteur", "Min", "Max", "Moy", "Écart-type", "Médiane"])
        for metric, s in stats.items():
            if s:
                ws_stat.append([
                    metric.replace("_", " ").title(),
                    s.get("min"), s.get("max"), s.get("mean"), s.get("std"), s.get("median"),
                ])
        for col in ["A","B","C","D","E","F"]:
            ws_stat.column_dimensions[col].width = 16

        # Sheet 4: AI Report
        ws_ai = wb.create_sheet("Rapport IA")
        ws_ai["A1"] = "ORBIT AI — Rapport de Diagnostic"
        ws_ai["A1"].font = Font(bold=True, size=14)
        ws_ai["A2"] = f"Généré le : {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
        ws_ai["A3"] = f"Fichier : {report.filename}"
        ws_ai["A5"] = report.ai_narrative or "Aucun rapport IA généré."
        ws_ai["A5"].alignment = Alignment(wrap_text=True)
        ws_ai.column_dimensions["A"].width = 100
        ws_ai.row_dimensions[5].height = 400

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── PDF (professional industrial report) ──────────────────────────────

    def to_pdf(self, report: Any) -> bytes:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            HRFlowable, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
        )

        BLU     = colors.HexColor("#1e40af")
        LBL     = colors.HexColor("#dbeafe")
        BLU2    = colors.HexColor("#1e3a8a")
        STRIPE  = colors.HexColor("#f0f4ff")
        RED     = colors.HexColor("#dc2626")
        ORANGE  = colors.HexColor("#ea580c")
        YELLOW  = colors.HexColor("#ca8a04")
        GREEN   = colors.HexColor("#16a34a")

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4,
                                leftMargin=2*cm, rightMargin=2*cm,
                                topMargin=2*cm, bottomMargin=2*cm)

        styles = getSampleStyleSheet()

        def sty(name, **kw):
            return ParagraphStyle(name, parent=styles["Normal"], **kw)

        title_sty   = sty("T",  fontSize=20, textColor=BLU, fontName="Helvetica-Bold", spaceAfter=4)
        h2_sty      = sty("H2", fontSize=12, textColor=BLU2, fontName="Helvetica-Bold", spaceBefore=14, spaceAfter=4)
        h3_sty      = sty("H3", fontSize=10, textColor=BLU2, fontName="Helvetica-Bold", spaceBefore=8, spaceAfter=2)
        body_sty    = sty("B",  fontSize=9,  leading=14)
        caption_sty = sty("C",  fontSize=8,  textColor=colors.grey)
        small_sty   = sty("S",  fontSize=8,  leading=12, textColor=colors.HexColor("#374151"))
        bold_sty    = sty("Bo", fontSize=9,  fontName="Helvetica-Bold")

        story: list = []

        analysis    = json.loads(report.analysis_json) if report.analysis_json else {}
        risk        = analysis.get("risk", {})
        anomaly     = analysis.get("anomaly", {})
        trends      = analysis.get("trends", {})
        stats       = analysis.get("statistics", {})
        xai         = analysis.get("xai", [])
        timeline    = analysis.get("health_timeline", [])
        rul         = analysis.get("rul", {})
        rf          = analysis.get("risk_factors", [])
        recs_prio   = analysis.get("recommendations_prioritized", [])
        mp          = analysis.get("motor_profile") or {}
        fault       = analysis.get("fault", "No Fault")
        health      = analysis.get("health_score", 0)
        conf        = analysis.get("confidence", 0)
        sev         = analysis.get("severity", "LOW")
        iso_zone    = analysis.get("iso_zone", "—")

        fault_fr    = _FAULT_FR.get(fault, fault)
        sev_fr      = _SEV_FR.get(sev, sev)
        sev_color   = {"CRITICAL": RED, "HIGH": ORANGE, "MEDIUM": YELLOW, "LOW": GREEN}.get(sev, GREEN)
        health_color = RED if health < 45 else ORANGE if health < 60 else YELLOW if health < 75 else GREEN

        # ── Cover header ────────────────────────────────────────────────
        story.append(Paragraph("⚙ ORBIT AI", sty("Logo", fontSize=11, textColor=colors.white,
                                                   fontName="Helvetica-Bold")))
        # Blue banner
        banner = Table([["ORBIT AI — Rapport de Diagnostic Industriel"]], colWidths=[17*cm])
        banner.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), BLU),
            ("TEXTCOLOR",  (0,0), (-1,-1), colors.white),
            ("FONTNAME",   (0,0), (-1,-1), "Helvetica-Bold"),
            ("FONTSIZE",   (0,0), (-1,-1), 16),
            ("TOPPADDING", (0,0), (-1,-1), 12),
            ("BOTTOMPADDING", (0,0), (-1,-1), 12),
            ("LEFTPADDING",   (0,0), (-1,-1), 10),
        ]))
        story.append(banner)
        story.append(Spacer(1, 0.3*cm))
        story.append(Paragraph(f"Fichier : {report.filename}  |  "
                               f"Généré le {datetime.utcnow().strftime('%d/%m/%Y %H:%M')} UTC  |  "
                               f"Lignes : {report.row_count:,}  |  Qualité : {report.quality_score:.1f}%",
                               caption_sty))
        story.append(Spacer(1, 0.4*cm))

        # ── Motor profile ────────────────────────────────────────────────
        if mp and any(mp.get(k) for k in ("name","manufacturer","nominal_power_kw","nominal_current_a")):
            story.append(Paragraph("Profil Moteur", h2_sty))
            mp_rows = [["Paramètre", "Valeur"]]
            for label, key, unit in [
                ("Nom / ID",          "name",              ""),
                ("Fabricant",         "manufacturer",      ""),
                ("Puissance nominale","nominal_power_kw",  " kW"),
                ("Tension nominale",  "nominal_voltage_v", " V"),
                ("Courant nominal",   "nominal_current_a", " A"),
                ("Vitesse nominale",  "nominal_speed_rpm", " tr/min"),
                ("Classe isolation",  "insulation_class",  ""),
                ("Classe efficacité", "efficiency_class",  ""),
                ("Indice protection", "protection_class",  ""),
            ]:
                val = mp.get(key)
                if val is not None:
                    mp_rows.append([label, f"{val}{unit}"])
            if len(mp_rows) > 1:
                story.append(self._make_table(mp_rows, col_widths=[6*cm, 11*cm]))
            story.append(Spacer(1, 0.3*cm))

        # ── Executive summary ────────────────────────────────────────────
        story.append(Paragraph("Synthèse Exécutive", h2_sty))
        exec_data = [
            ["Score de Santé",   f"{health}/100  —  {_STATUS_FR.get(analysis.get('health_status',''), '—')}"],
            ["Défaut Détecté",   fault_fr],
            ["Sévérité",         sev_fr],
            ["Confiance ML",     f"{conf}%"],
            ["Zone ISO 10816",   iso_zone],
            ["RUL estimé",       rul.get("value", "—") + f"  (confiance {rul.get('confidence','—')})"],
            ["Risque 7 jours",   f"{risk.get('days_7', 0):.1f}%"],
            ["Risque 30 jours",  f"{risk.get('days_30', 0):.1f}%"],
            ["Anomalies",        f"{anomaly.get('count',0)} lectures ({anomaly.get('percentage',0):.1f}%)"],
        ]
        t_exec = self._make_table(exec_data, col_widths=[5.5*cm, 11.5*cm])
        # Color the severity row
        t_exec.setStyle(TableStyle([
            ("TEXTCOLOR", (1, 1), (1, 1), health_color),
            ("FONTNAME",  (1, 1), (1, 1), "Helvetica-Bold"),
            ("TEXTCOLOR", (1, 3), (1, 3), sev_color),
            ("FONTNAME",  (1, 3), (1, 3), "Helvetica-Bold"),
        ]))
        story.append(t_exec)
        story.append(Spacer(1, 0.4*cm))

        # ── Health Timeline ──────────────────────────────────────────────
        if timeline:
            story.append(Paragraph("Chronologie de Santé", h2_sty))
            tl_rows = [["Phase", "Santé moy.", "Portion du dataset"]]
            phase_icons = {"Sain":"✅","Dégradation Précoce":"⚠","Avertissement":"🔶","Critique":"🚨"}
            for ph in timeline:
                icon = phase_icons.get(ph["phase"], "•")
                tl_rows.append([
                    f"{icon}  {ph['phase']}",
                    f"{ph['avg_health']}/100",
                    f"{ph['start_pct']}% → {ph['end_pct']}%",
                ])
            story.append(self._make_table(tl_rows, col_widths=[7*cm, 4*cm, 6*cm], has_header=True))
            story.append(Spacer(1, 0.4*cm))

        # ── XAI Feature contributions ────────────────────────────────────
        if xai:
            story.append(Paragraph(f"Analyse Explicable (XAI) — Diagnostic : {fault_fr}", h2_sty))
            story.append(Paragraph("Facteurs ayant contribué à ce diagnostic (poids normalisés) :", small_sty))
            story.append(Spacer(1, 0.2*cm))
            xai_rows = [["Facteur", "Contribution", "Visualisation"]]
            for item in xai[:7]:
                pct  = item["contribution"]
                bar  = "█" * int(pct / 5) + "░" * (20 - int(pct / 5))
                xai_rows.append([item["feature"], f"{pct}%", bar])
            story.append(self._make_table(xai_rows, col_widths=[7*cm, 2.5*cm, 7.5*cm], has_header=True))
            story.append(Spacer(1, 0.4*cm))

        # ── Risk factors ─────────────────────────────────────────────────
        if rf:
            story.append(Paragraph("Facteurs de Risque", h2_sty))
            for r_factor in rf:
                story.append(Paragraph(f"⚠  {r_factor}", small_sty))
            story.append(Spacer(1, 0.3*cm))

        # ── Sensor statistics ────────────────────────────────────────────
        story.append(Paragraph("Statistiques des Capteurs", h2_sty))
        stat_rows = [["Capteur", "Min", "Max", "Moy", "Écart-type", "Zone ISO"]]
        iso_info = {"vibration": iso_zone}
        label_map = {
            "temperature":"Température (°C)","vibration":"Vibration (mm/s)",
            "current":"Courant (A)","voltage":"Tension (V)","power":"Puissance (kW)",
            "load":"Charge (ratio)","power_factor":"Cos φ","thd":"THD (%)",
        }
        for key, s in stats.items():
            if s and key != "health_score":
                row = [
                    label_map.get(key, key.replace("_"," ").title()),
                    f"{s.get('min',0):.2f}", f"{s.get('max',0):.2f}",
                    f"{s.get('mean',0):.2f}", f"{s.get('std',0):.2f}",
                    iso_info.get(key, ""),
                ]
                stat_rows.append(row)
        story.append(self._make_table(stat_rows,
                     col_widths=[5.5*cm, 2*cm, 2*cm, 2.5*cm, 2.5*cm, 2.5*cm], has_header=True))
        story.append(Spacer(1, 0.4*cm))

        # ── Sensor trends ─────────────────────────────────────────────────
        if trends:
            story.append(Paragraph("Tendances Capteurs", h2_sty))
            tr_rows = [["Capteur", "Tendance"]]
            for k, v in trends.items():
                tr_rows.append([k.replace("_"," ").title(), _TREND_FR.get(v, v)])
            story.append(self._make_table(tr_rows, col_widths=[7*cm, 10*cm], has_header=True))
            story.append(Spacer(1, 0.4*cm))

        # ── Multi-Agent (run on-demand) ───────────────────────────────────
        try:
            from app.services.multi_agent_service import MultiAgentService
            ma_result  = MultiAgentService().analyze(analysis)
            agents     = ma_result.get("agents", [])
            synthesis  = ma_result.get("synthesis", {})
            if agents:
                story.append(Paragraph("Analyse Multi-Agent IA", h2_sty))
                consensus_rows = [["Agent", "Domaine", "Sévérité", "Confiance"]]
                sev_map = {"LOW":"FAIBLE","MEDIUM":"MOYEN","HIGH":"ÉLEVÉ","CRITICAL":"CRITIQUE"}
                for ag in agents:
                    consensus_rows.append([
                        ag["title"], ag["domain"][:40],
                        sev_map.get(ag["severity"], ag["severity"]),
                        f"{min(100,ag['confidence'])}%",
                    ])
                consensus_rows.append([
                    "🎯 Coordinateur", "Synthèse globale",
                    sev_map.get(synthesis.get("overall_severity",""), synthesis.get("overall_severity","")),
                    f"{synthesis.get('consensus_confidence',0)}%",
                ])
                story.append(self._make_table(consensus_rows,
                             col_widths=[4.5*cm, 6.5*cm, 3*cm, 3*cm], has_header=True))
                verdict = synthesis.get("verdict","")
                if verdict:
                    story.append(Spacer(1, 0.2*cm))
                    story.append(Paragraph(f"Verdict : {verdict}  |  Délai : {synthesis.get('timeline','—')}", bold_sty))
                story.append(Spacer(1, 0.3*cm))
        except Exception:
            pass

        # ── Maintenance plan ──────────────────────────────────────────────
        if recs_prio:
            story.append(Paragraph("Plan de Maintenance Priorisé", h2_sty))
            urgency_fr = {"immediate":"⚡ Immédiat","days":"📅 Cette semaine","weeks":"🗓 Ce mois","months":"📆 Planifié"}
            prio_rows = [["Priorité", "Action", "Délai"]]
            for rec in recs_prio:
                prio_rows.append([
                    str(rec["priority"]),
                    rec["action"],
                    urgency_fr.get(rec.get("urgency",""), rec.get("urgency","")),
                ])
            story.append(self._make_table(prio_rows, col_widths=[2*cm, 11.5*cm, 3.5*cm], has_header=True))
            story.append(Spacer(1, 0.3*cm))

        # ── AI Narrative ──────────────────────────────────────────────────
        if report.ai_narrative:
            story.append(PageBreak())
            story.append(Paragraph("Rapport de Diagnostic IA (Qwen)", h2_sty))
            story.append(HRFlowable(width="100%", thickness=0.5, color=colors.lightgrey))
            story.append(Spacer(1, 0.2*cm))
            for para in report.ai_narrative.split("\n\n"):
                para = para.strip()
                if para:
                    if para == para.upper() and len(para) < 60:
                        story.append(Paragraph(para, h3_sty))
                    else:
                        story.append(Paragraph(para.replace("\n", " "), body_sty))
                    story.append(Spacer(1, 0.15*cm))

        # ── Standards reference ───────────────────────────────────────────
        story.append(Spacer(1, 0.5*cm))
        story.append(Paragraph("Normes et Références", h2_sty))
        norms = [
            ["ISO 10816 / ISO 20816", "Évaluation des vibrations des machines tournantes"],
            ["IEC 60034-1",           "Machines électriques tournantes — limites thermiques"],
            ["IEC 60034-30",          "Classes de rendement moteur IE1–IE4"],
            ["IEEE 112",              "Méthodes de mesure du rendement moteurs asynchrones"],
            ["IEEE 519",              "Limites de distorsion harmonique (THD)"],
            ["ISO 1940-1",            "Équilibrage — grades G (G2.5 pour moteurs industriels)"],
            ["ISO 13373",             "Surveillance de l'état des machines — vibrations"],
        ]
        story.append(self._make_table(norms, col_widths=[4*cm, 13*cm], has_header=True))

        # ── Footer ────────────────────────────────────────────────────────
        story.append(Spacer(1, 0.5*cm))
        story.append(HRFlowable(width="100%", thickness=0.5, color=BLU))
        story.append(Paragraph(
            f"ORBIT AI Industrial Copilot — Plateforme de Maintenance Prédictive  |  "
            f"Rapport généré le {datetime.utcnow().strftime('%d/%m/%Y à %H:%M')} UTC",
            caption_sty,
        ))

        doc.build(story)
        return buf.getvalue()

    # ── Helpers ──────────────────────────────────────────────────────────

    def _make_table(
        self,
        rows: list[list],
        col_widths: list | None = None,
        has_header: bool = False,
    ) -> Any:
        from reportlab.lib import colors
        from reportlab.platypus import Table, TableStyle

        BLU    = colors.HexColor("#1e40af")
        STRIPE = colors.HexColor("#f0f4ff")
        LBL    = colors.HexColor("#dbeafe")

        t = Table(rows, colWidths=col_widths)
        style = [
            ("FONTSIZE",      (0, 0), (-1, -1), 9),
            ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#d1d5db")),
            ("ROWBACKGROUNDS",(0, 0), (-1, -1), [colors.white, STRIPE]),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING",   (0, 0), (-1, -1), 6),
        ]
        if has_header:
            style += [
                ("BACKGROUND",  (0, 0), (-1, 0), BLU),
                ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
                ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, STRIPE]),
            ]
        else:
            style += [
                ("BACKGROUND", (0, 0), (0, -1), LBL),
                ("FONTNAME",   (0, 0), (0, -1), "Helvetica-Bold"),
            ]
        t.setStyle(TableStyle(style))
        return t

    @staticmethod
    def _write_header_row(ws: Any, headers: list[str]) -> None:
        from openpyxl.styles import Font, PatternFill
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1e40af")

    # Legacy helpers (kept for compatibility)
    def _simple_table(self, data, col_widths=None):
        return self._make_table(data, col_widths=col_widths, has_header=False)

    def _table_with_header(self, rows):
        return self._make_table(rows, has_header=True)
